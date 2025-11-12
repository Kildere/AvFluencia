
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def auditar_excel_and_save_reports(file_path: str, sheet_name: str = "Consolidação"):
    """Audita a planilha: conta fórmulas e erros; salva relatórios na MESMA pasta do arquivo.
    Retorna dict com contagens e pasta.
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        # Se a aba não existir, retornar 0 e continuar
        return {'formulas': 0, 'erros': 0, 'folder': os.path.dirname(file_path)}

    ws = wb[sheet_name]
    registros = []
    formulas = 0
    erros = 0

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            val = str(cell.value).strip().upper() if cell.value else ""
            if val.startswith("="):
                formulas += 1
                registros.append({
                    "CELULA": f"{get_column_letter(cell.column)}{cell.row}",
                    "TIPO": "FÓRMULA",
                    "CONTEUDO": cell.value
                })
            elif val in ["#N/D", "#DIV/0!", "#REF!", "#VALOR!", "#NOME?"]:
                erros += 1
                registros.append({
                    "CELULA": f"{get_column_letter(cell.column)}{cell.row}",
                    "TIPO": "ERRO",
                    "CONTEUDO": val
                })

    df = pd.DataFrame(registros)
    folder = os.path.dirname(file_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    if not df.empty:
        xlsx_path = os.path.join(folder, f"relatorio_auditoria_{ts}.xlsx")
        csv_path = os.path.join(folder, f"relatorio_auditoria_{ts}.csv")
        try:
            df.to_excel(xlsx_path, index=False)
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        except Exception:
            # Mesmo que não consiga salvar (permissão), ainda retorna as contagens
            pass

    return {'formulas': formulas, 'erros': erros, 'folder': folder}
